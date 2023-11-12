import os
from openpyxl import Workbook, load_workbook

class UserSystem:
    def __init__(self, file_path, excel_path):
        self.file_path = file_path
        self.excel_path = excel_path
        self.users = {}
        self.load_users()

    def load_users(self):
        try:
            with open(self.file_path, 'r') as f:
                for line in f:
                    username, password = line.strip().split(',')
                    self.users[username] = password
        except FileNotFoundError:
            print(f"{self.file_path} not found. Creating a new file.")
            self.save_users()

        try:
            workbook = load_workbook(self.excel_path)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                username, password = row
                self.users[username] = password
        except FileNotFoundError:
            print(f"{self.excel_path} not found. Creating a new file.")
            self.save_users()

    def save_users(self):
        with open(self.file_path, 'w') as f:
            for username, password in self.users.items():
                f.write(f"{username},{password}\n")

        workbook = Workbook()
        sheet = workbook.active
        for username, password in self.users.items():
            sheet.append((username, password))
        workbook.save(self.excel_path)

    def register(self, username, password):
        if username in self.users:
            print("Username already exists. Please choose a different username.")
        else:
            self.users[username] = password
            self.save_users()
            print("Registration successful.")

    def login(self, username, password):
        if username not in self.users:
            print("Username not found. Please register first.")
        elif self.users[username] != password:
            print("Incorrect password. Please try again.")
        else:
            print("Login successful.")

# Example usage:
file_path = "users.txt"
excel_path = "users.xlsx"
user_system = UserSystem(file_path, excel_path)

# Registering users
user_system.register("user1", "password1")
user_system.register("user2", "password2")

# Logging in users
user_system.login("user1", "password1")
user_system.login("user2", "password3")  # Wrong password
user_system.login("user3", "password3")  # User not registered






from openpyxl import load_workbook

class UserSystem:
    def __init__(self, filename):
        self.filename = filename

    def load_users(self):
        try:
            workbook = load_workbook(self.filename)
            sheet = workbook.active
            users = {}
            for row in sheet.iter_rows(values_only=True):
                username, password = row
                users[username] = password
            return users
        except FileNotFoundError:
            print(f"{self.filename} not found. Please create the file.")
            return None

    def register(self, username, password):
        users = self.load_users()
        if users is None:
            return 

        if username in users:
            print("Username already exists. Please choose a different username.")
        else:
            users[username] = password
            self.save_users(users)
            print("Registration successful.")

    def save_users(self, users):
        workbook = Workbook()
        sheet = workbook.active
        for username, password in users.items():
            sheet.append((username, password))
        workbook.save(self.filename)

    def login(self, username, password):
        users = self.load_users()
        if users is None:
            return

        if username not in users:
            print("Username not found. Please register first.")
        elif users[username] != password:
            print("Incorrect password. Please try again.")
        else:
            print("Login successful.")

# Example usage:
user_system = UserSystem("users.xlsx")

# Registering users
user_system.register("user1", "password1")
user_system.register("user2", "password2")

# Logging in users
user_system.login("user1", "password1")
user_system.login("user2", "password3")  # Wrong password
user_system.login("user3", "password3")  # User not registered
